"""
BLOQUE 1 — Estandarización y validación de coordenadas
Protocolo SiB Colombia / Instituto Humboldt
© 2026 Ximena Bedoya Araque · Universidad CES · Colecciones Biológicas CBUCES

Reglas geográficas Colombia (CCO / DANE):
  Latitud continental:  -4 a 12  (1 dígito antes del punto)
  Longitud continental: -82 a -66 (2 dígitos antes del punto, siempre negativo)
  Zona marítima: lat hasta 16, lon hasta -82 → marcar para verificación
"""

import pandas as pd
import numpy as np
import re
import unicodedata


def normalizar(texto):
    if pd.isna(texto): return ""
    t = unicodedata.normalize("NFD", str(texto).strip())
    return "".join(c for c in t if unicodedata.category(c) != "Mn").lower()


def limpiar_valor_coordenada(valor):
    """
    Limpia caracteres inválidos de una coordenada.
    Maneja todos los casos reales encontrados en la base:
      602825         → entero sin punto
      6.787.898      → dos puntos → quitar el primero
      6.913.160.858  → tres puntos → quitar todos menos el último
      ?74,94166      → signo ? + coma → limpiar
      6.14°          → símbolo grado → quitar
      06°04'47.3''   → GMS → dejar para conversión
      6°14'02.4''N   → GMD con letra → dejar para conversión
    """
    if pd.isna(valor): return None
    s = str(valor).strip()
    if s in ("", "nan"): return None

    # Quitar signo de interrogación
    s = s.replace("?", "").strip()

    # Si tiene formato GMS (grados minutos segundos) → devolver tal cual
    if "''" in s or (re.search(r"[°']", s) and re.search(r"[NSEWО]", s, re.IGNORECASE)):
        return s

    # Si tiene símbolo de grado sin segundos (GMD)
    if "°" in s:
        return s

    # Reemplazar coma decimal por punto
    s = s.replace(",", ".")

    # Múltiples puntos → quitar todos menos el del decimal
    # "6.787.898" → "6787898" y luego se inserta el punto en la posición correcta
    # "6.913.160.858" → "6913160858"
    if s.count(".") > 1:
        # Quitar todos los puntos — el punto decimal se infiere después
        negativo = s.startswith("-")
        s_limpio = s.replace("-", "").replace(".", "")
        s = ("-" if negativo else "") + s_limpio

    # Quitar letras N/S/E/W/O si quedaron sueltas
    s = re.sub(r"[NSEWOnsewo]$", "", s).strip()
    s = re.sub(r"^[NSEWOnsewo]", "", s).strip()

    # Quitar símbolo de grado si quedó
    s = s.replace("°", "").strip()

    if s in ("", "-"): return None
    return s


def insertar_punto_decimal(valor_str, es_lon=False):
    """
    Inserta punto decimal en posición correcta según rango Colombia.
    Latitud:  1 dígito antes del punto → rango -5 a 14
    Longitud: 2 dígitos antes del punto → rango -82 a -66
    """
    s = str(valor_str).strip()
    negativo = s.startswith("-")
    s_abs = s.lstrip("-")

    for pos in range(1, min(len(s_abs), 4)):
        try:
            candidato = float(s_abs[:pos] + "." + s_abs[pos:])
            if negativo: candidato = -candidato
            if es_lon:
                if -82 <= candidato <= -66 or 66 <= candidato <= 82:
                    return round(-abs(candidato), 6)
            else:
                if -5 <= candidato <= 14:
                    return round(candidato, 6)
        except:
            pass
    return None


def convertir_gms(texto):
    """GMS/GMD → decimal. Acepta: 06°04'47.3'' / 6°14.158'N / 75°40'21.1''W"""
    texto = str(texto).strip()
    negativo = any(c in texto.upper() for c in ["S", "W", "O"])
    nums = re.findall(r"[\d]+[.,]?[\d]*", texto)
    if not nums: return None
    try:
        if len(nums) >= 3:
            d = (float(nums[0]) +
                 float(nums[1].replace(",", ".")) / 60 +
                 float(nums[2].replace(",", ".")) / 3600)
        elif len(nums) == 2:
            d = float(nums[0]) + float(nums[1].replace(",", ".")) / 60
        else:
            d = float(nums[0].replace(",", "."))
        return -d if negativo else d
    except:
        return None


def convertir_a_decimal(valor_limpio, es_lon=False):
    """
    Convierte un valor ya limpio a decimal WGS84.
    Retorna: (valor_decimal, fuente_str) o (None, mensaje_error)
    """
    if valor_limpio is None:
        return None, "sin dato"

    s = str(valor_limpio).strip()

    # GMS
    if "''" in s:
        d = convertir_gms(s)
        if d is not None:
            if es_lon and d > 0: d = -d
            if es_lon and not (-82 <= d <= -66): return None, "GMS fuera de rango Colombia"
            if not es_lon and not (-5 <= d <= 14): return None, "GMS fuera de rango Colombia"
            return round(d, 6), "GMS convertido"
        return None, "error conversión GMS"

    # GMD
    if "°" in s:
        d = convertir_gms(s)
        if d is not None:
            if es_lon and d > 0: d = -d
            if es_lon and not (-82 <= d <= -66): return None, "GMD fuera de rango Colombia"
            if not es_lon and not (-5 <= d <= 14): return None, "GMD fuera de rango Colombia"
            return round(d, 6), "GMD convertido"
        return None, "error conversión GMD"

    # Entero sin punto
    if re.match(r"^-?\d{5,10}$", s):
        d = insertar_punto_decimal(s, es_lon)
        if d is not None:
            return d, "entero → decimal (punto inferido)"
        return None, "entero sin punto: no se pudo inferir posición del punto"

    # Decimal directo
    try:
        d = float(s)
        # Corregir longitud positiva en rango Colombia
        if es_lon and 66 <= d <= 82:
            d = -d
        if es_lon and not (-82 <= d <= -66):
            return None, f"longitud fuera de rango Colombia: {d}"
        if not es_lon and not (-5 <= d <= 14):
            return None, f"latitud fuera de rango Colombia: {d}"
        return round(d, 6), "decimal directo"
    except:
        pass

    return None, f"formato no reconocido: {repr(s)}"


def validar_rango_colombia(lat, lon):
    """
    Valida si el punto está en Colombia.
    Retorna: ('continental', 'maritima', 'fuera')
    """
    if lat is None or lon is None:
        return "sin_coordenadas"
    if -4 <= lat <= 12 and -79 <= lon <= -66:
        return "continental"
    if -5 <= lat <= 16 and -82 <= lon <= -60:
        return "maritima"
    return "fuera"


def aplicar_bloque1(df84_path, df180_path, idioma="es"):
    """
    Lee los dos archivos, los une y limpia todas las coordenadas.

    Retorna DataFrame unido con columnas nuevas:
      lat_decimal_calculada  → latitud WGS84 corregida
      lon_decimal_calculada  → longitud WGS84 corregida
      formato_coordenada     → tipo detectado (GMS, GMD, decimal, entero)
      conversion_estado      → OK / advertencia / error
      conversion_nota        → explicación
      rango_colombia         → continental / maritima / fuera / sin_coordenadas
    """
    # Leer archivos
    df180 = pd.read_excel(df180_path, header=0)
    df84  = pd.read_excel(df84_path,  header=0)

    # Normalizar nombre columna localidad estandarizada
    df84  = df84.rename(columns={"*Localidad Estandarizada": "*Localidad estandarizada"})

    # Agregar columna Origen
    df180["Origen"] = "Con coordenadas"
    df84["Origen"]  = "Sin coordenadas"

    # Unir — columnas comunes
    cols_comunes = [c for c in df180.columns if c in df84.columns]
    df = pd.concat(
        [df180[cols_comunes], df84[cols_comunes]],
        ignore_index=True
    )

    print(f"  Total registros unidos: {len(df)}")
    print(f"    Con coordenadas: {(df['Origen']=='Con coordenadas').sum()}")
    print(f"    Sin coordenadas: {(df['Origen']=='Sin coordenadas').sum()}")

    # Detectar columnas de coordenadas según idioma
    if idioma == "es":
        col_lat = "Latitud original"
        col_lon = "Longitud original"
    else:
        col_lat = "verbatimLatitude"
        col_lon = "verbatimLongitude"

    # Inicializar columnas de resultado
    df["lat_decimal_calculada"] = np.nan
    df["lon_decimal_calculada"] = np.nan
    df["formato_coordenada"]    = ""
    df["conversion_estado"]     = ""
    df["conversion_nota"]       = ""
    df["rango_colombia"]        = ""

    procesados = errores = sin_dato = 0

    for idx in df.index:
        lat_raw = df.at[idx, col_lat]
        lon_raw = df.at[idx, col_lon]

        # Sin dato
        if (pd.isna(lat_raw) or str(lat_raw).strip() in ("", "nan")) and \
           (pd.isna(lon_raw) or str(lon_raw).strip() in ("", "nan")):
            df.at[idx, "conversion_estado"] = "sin coordenadas"
            df.at[idx, "rango_colombia"]    = "sin_coordenadas"
            sin_dato += 1
            continue

        # Limpiar
        lat_limpia = limpiar_valor_coordenada(lat_raw)
        lon_limpia = limpiar_valor_coordenada(lon_raw)

        # Detectar formato
        if lat_limpia and "''" in str(lat_limpia):
            formato = "GMS"
        elif lat_limpia and "°" in str(lat_limpia):
            formato = "GMD"
        elif lat_limpia and re.match(r"^-?\d{5,10}$", str(lat_limpia)):
            formato = "entero sin punto"
        else:
            formato = "decimal"
        df.at[idx, "formato_coordenada"] = formato

        # Convertir
        lat_dec, lat_nota = convertir_a_decimal(lat_limpia, es_lon=False)
        lon_dec, lon_nota = convertir_a_decimal(lon_limpia, es_lon=True)

        if lat_dec is not None and lon_dec is not None:
            df.at[idx, "lat_decimal_calculada"] = lat_dec
            df.at[idx, "lon_decimal_calculada"] = lon_dec
            rango = validar_rango_colombia(lat_dec, lon_dec)
            df.at[idx, "rango_colombia"] = rango

            if rango == "continental":
                df.at[idx, "conversion_estado"] = "OK"
                df.at[idx, "conversion_nota"]   = lat_nota
            elif rango == "maritima":
                df.at[idx, "conversion_estado"] = "Revisar"
                df.at[idx, "conversion_nota"]   = (
                    "Zona maritima — verificar con fuente oficial CCO: "
                    "https://cco.gov.co/comision/mapa-esquematico/"
                )
            else:
                df.at[idx, "conversion_estado"] = "Error"
                df.at[idx, "conversion_nota"]   = (
                    f"Fuera de Colombia: lat={lat_dec}, lon={lon_dec}"
                )
            procesados += 1
        else:
            df.at[idx, "conversion_estado"] = "Revisar"
            df.at[idx, "conversion_nota"]   = (
                f"lat: {lat_nota} | lon: {lon_nota} | "
                f"valor original: {lat_raw} / {lon_raw}"
            )
            df.at[idx, "rango_colombia"] = "no_convertido"
            errores += 1

    print(f"\n  Resultados Bloque 1:")
    print(f"    Convertidos OK:        {procesados}")
    print(f"    Requieren revision:    {errores}")
    print(f"    Sin coordenadas:       {sin_dato}")
    print(f"\n  Distribución rango Colombia:")
    print(df["rango_colombia"].value_counts().to_string())

    return df


# ─────────────────────────────────────────────────────────────────
# BLOQUE 4 — Incertidumbre (Tabla 2 del manual)
# Separado aquí para importar fácilmente desde otros bloques
# ─────────────────────────────────────────────────────────────────

# Tabla 2 del manual — valores exactos
INCERTIDUMBRE_COORDENADAS = {
    # GMS (grados minutos segundos)
    "gms_1_grado":     156904,
    "gms_1_minuto":    2615,
    "gms_1_segundo":   44,
    "gms_01_segundo":  5,
    "gms_001_segundo": 1,
    # GMD (grados minutos decimales)
    "gmd_01_minuto":   262,
    "gmd_001_minuto":  27,
    "gmd_0001_minuto": 3,
    # Decimal
    "dec_01":    15691,
    "dec_001":   1570,
    "dec_0001":  157,
    "dec_00001": 16,
    "dec_6dec":  2,     # enteros sin punto → 5-6 decimales → 2m
}


def incertidumbre_por_formato(lat_original, formato):
    """
    Retorna incertidumbre en metros según Tabla 2 del manual.
    """
    if pd.isna(lat_original): return None
    s = str(lat_original).strip()

    # GMS
    if formato == "GMS":
        partes = re.findall(r"\d+[.,]?\d*", s)
        if len(partes) >= 3:
            seg = partes[2]
            if "." in seg or "," in seg:
                decimales = len(seg.split(".")[-1]) if "." in seg else len(seg.split(",")[-1])
                if decimales >= 2: return INCERTIDUMBRE_COORDENADAS["gms_001_segundo"]
                return INCERTIDUMBRE_COORDENADAS["gms_01_segundo"]
            return INCERTIDUMBRE_COORDENADAS["gms_1_segundo"]
        return INCERTIDUMBRE_COORDENADAS["gms_1_minuto"]

    # GMD
    if formato == "GMD":
        partes = re.findall(r"\d+[.,]?\d*", s)
        if len(partes) >= 2:
            min_str = partes[1]
            if "." in min_str or "," in min_str:
                decimales = len(min_str.split(".")[-1]) if "." in min_str else len(min_str.split(",")[-1])
                if decimales >= 3: return INCERTIDUMBRE_COORDENADAS["gmd_0001_minuto"]
                if decimales == 2: return INCERTIDUMBRE_COORDENADAS["gmd_001_minuto"]
                return INCERTIDUMBRE_COORDENADAS["gmd_01_minuto"]
        return INCERTIDUMBRE_COORDENADAS["gmd_01_minuto"]

    # Entero sin punto → se infiere 5-6 decimales → 2m
    if formato == "entero sin punto":
        return INCERTIDUMBRE_COORDENADAS["dec_6dec"]

    # Decimal directo
    if formato == "decimal":
        try:
            val = float(s.replace(",", "."))
            s_dec = str(abs(val))
            if "." in s_dec:
                decimales = len(s_dec.split(".")[1].rstrip("0"))
                if decimales >= 5: return INCERTIDUMBRE_COORDENADAS["dec_6dec"]
                if decimales == 4: return INCERTIDUMBRE_COORDENADAS["dec_00001"]
                if decimales == 3: return INCERTIDUMBRE_COORDENADAS["dec_0001"]
                if decimales == 2: return INCERTIDUMBRE_COORDENADAS["dec_001"]
                if decimales == 1: return INCERTIDUMBRE_COORDENADAS["dec_01"]
        except:
            pass

    return None


def calcular_incertidumbre_total(row, gdf_gadm=None, idioma="es"):
    """
    Incertidumbre total = datum + coordenadas + extensión (radio máximo municipio)
    Solo para registros con coordenadas (Nivel 1).
    Crea columna nueva 'Incertidumbre de coordenadas (m)' — NO sobreescribe 'Precisión (m)'.
    """
    col_datum   = "Datum" if idioma == "es" else "geodeticDatum"
    col_lat_o   = "Latitud original" if idioma == "es" else "verbatimLatitude"
    formato     = str(row.get("formato_coordenada", ""))

    # Datum
    datum = str(row.get(col_datum, "")).strip()
    if datum in ("", "nan", "WGS 84 (asumido)"):
        inc_datum = 500
    else:
        inc_datum = 0

    # Coordenadas
    inc_coord = incertidumbre_por_formato(row.get(col_lat_o), formato) or 0

    # Extensión (radio máximo municipio desde GADM)
    inc_extension = 0
    if gdf_gadm is not None:
        try:
            import unicodedata as ud
            muni_norm = "".join(
                c for c in ud.normalize("NFD", str(row.get("*Municipio","")).strip().lower())
                if ud.category(c) != "Mn"
            )
            hits = gdf_gadm[gdf_gadm["muni_norm"] == muni_norm]
            if not hits.empty:
                geom_proj = hits.iloc[0].geometry
                from shapely.ops import transform
                import pyproj
                proj = pyproj.Transformer.from_crs("EPSG:4326","EPSG:3116",always_xy=True).transform
                geom_m = transform(proj, geom_proj)
                inc_extension = int(geom_m.minimum_rotated_rectangle.length / 2)
        except:
            pass

    total = inc_datum + inc_coord + inc_extension
    return total if total > 0 else None

"""
BLOQUE 2 — Validación municipio (shapefile GADM)
Protocolo SiB Colombia / Instituto Humboldt
© 2026 Ximena Bedoya Araque · Universidad CES · Colecciones Biológicas CBUCES

Verifica si el punto (lat, lon) cae dentro del municipio reportado.
Fuente: GADM v4.1 — gadm41_COL_2.json
Resultados:
  ✅ Dentro del municipio reportado
  ⚠  Municipio vecino — indica cuál
  ❌ Fuera de Colombia
"""

import pandas as pd
import geopandas as gpd
import unicodedata
from shapely.geometry import Point


def normalizar(texto):
    if pd.isna(texto): return ""
    t = unicodedata.normalize("NFD", str(texto).strip())
    return "".join(c for c in t if unicodedata.category(c) != "Mn").lower()


def cargar_gadm(ruta_gadm):
    gdf = gpd.read_file(ruta_gadm)
    gdf["depto_norm"] = gdf["NAME_1"].apply(normalizar)
    gdf["muni_norm"]  = gdf["NAME_2"].apply(normalizar)
    if gdf.crs is None or gdf.crs.to_epsg() != 4326:
        gdf = gdf.set_crs("EPSG:4326")
    return gdf


def validar_punto_municipio(lat, lon, municipio_rep, depto_rep, gdf_gadm):
    """
    Retorna (estado, municipio_detectado, depto_detectado, mensaje)
    estado: 'OK' | 'Revisar' | 'Error'
    """
    if lat is None or lon is None:
        return "Error", "", "", "Sin coordenadas convertibles"

    if not (-5 <= lat <= 14 and -82 <= lon <= -66):
        if -5 <= lat <= 16 and -84 <= lon <= -60:
            return "Revisar", "", "", "Zona maritima — verificar con CCO"
        return "Error", "", "", f"Fuera de Colombia: lat={lat}, lon={lon}"

    punto = Point(lon, lat)
    hits  = gdf_gadm[gdf_gadm.geometry.contains(punto)]
    if hits.empty:
        hits = gdf_gadm[gdf_gadm.geometry.distance(punto) < 0.01]
    if hits.empty:
        return "Error", "", "", "Punto fuera de todos los municipios de Colombia"

    muni_det  = hits.iloc[0]["NAME_2"]
    depto_det = hits.iloc[0]["NAME_1"]

    if normalizar(municipio_rep) == normalizar(muni_det):
        return "OK", muni_det, depto_det, f"Dentro de {muni_det}, {depto_det}"
    if normalizar(depto_rep) == normalizar(depto_det):
        return "Revisar", muni_det, depto_det, \
            f"Reportado: {municipio_rep} | Detectado: {muni_det} ({depto_det})"
    return "Error", muni_det, depto_det, \
        f"Reportado: {municipio_rep} ({depto_rep}) | Detectado: {muni_det} ({depto_det})"


def aplicar_bloque2(df, ruta_gadm, idioma="es"):
    if idioma == "es":
        col_lat   = "lat_decimal_calculada"
        col_lon   = "lon_decimal_calculada"
        col_muni  = "*Municipio"
        col_depto = "*Departamento"
    else:
        col_lat   = "lat_decimal_calculada"
        col_lon   = "lon_decimal_calculada"
        col_muni  = "county"
        col_depto = "stateProvince"

    print("  Cargando GADM Colombia...")
    gdf_gadm = cargar_gadm(ruta_gadm)
    print(f"  {len(gdf_gadm)} municipios cargados")

    df = df.copy()
    df["validacion_b2"]       = ""
    df["municipio_detectado"] = ""
    df["depto_detectado"]     = ""
    df["mensaje_b2"]          = ""

    con_coord = df[df["lat_decimal_calculada"].notna()].index
    print(f"  Validando {len(con_coord)} registros con coordenadas...")

    for idx in con_coord:
        row    = df.loc[idx]
        lat    = row.get(col_lat)
        lon    = row.get(col_lon)
        muni   = row.get(col_muni, "")
        depto  = row.get(col_depto, "")

        estado, muni_d, depto_d, msg = validar_punto_municipio(
            lat, lon, muni, depto, gdf_gadm)

        df.at[idx, "validacion_b2"]       = estado
        df.at[idx, "municipio_detectado"] = muni_d
        df.at[idx, "depto_detectado"]     = depto_d
        df.at[idx, "mensaje_b2"]          = msg

    print("\n  Resultados Bloque 2:")
    print(df["validacion_b2"].value_counts().to_string())
    return df
"""
BLOQUE 3 — Elevación vía API
Protocolo SiB Colombia / Instituto Humboldt
© 2026 Ximena Bedoya Araque · Universidad CES · Colecciones Biológicas CBUCES

Solo se ejecuta si el Bloque 2 no marcó Error.
API principal: Open-Elevation
API respaldo:  Open-Topo-Data
Diferencia <= 100m → usar API automáticamente ✅
Diferencia >  100m → marcar ⚠ curador decide
"""

import pandas as pd
import requests
import time


def obtener_elevacion_api(lat, lon):
    """
    Intenta Open-Elevation primero, Open-Topo-Data como respaldo.
    Retorna (elevacion_metros, fuente) o (None, mensaje_error)
    """
    # Open-Elevation
    try:
        url = f"https://api.open-elevation.com/api/v1/lookup?locations={lat},{lon}"
        r = requests.get(url, timeout=8)
        if r.status_code == 200:
            elev = r.json()["results"][0]["elevation"]
            return elev, "Open-Elevation"
    except Exception:
        pass

    # Open-Topo-Data (respaldo)
    try:
        url2 = f"https://api.opentopodata.org/v1/srtm90m?locations={lat},{lon}"
        r2 = requests.get(url2, timeout=8)
        if r2.status_code == 200:
            elev = r2.json()["results"][0]["elevation"]
            return elev, "Open-Topo-Data"
    except Exception:
        pass

    return None, "API no disponible"


def aplicar_bloque3(df, idioma="es"):
    if idioma == "es":
        col_elev_min = "Elevación mínima (msnm)"
        col_elev_max = "Elevación máxima (msnm)"
    else:
        col_elev_min = "minimumElevationInMeters"
        col_elev_max = "maximumElevationInMeters"

    df = df.copy()
    df["elevacion_api"]      = None
    df["elevacion_fuente"]   = ""
    df["elevacion_estado"]   = ""
    df["elevacion_nota"]     = ""

    # Solo procesar si B2 no es Error
    mask = (
        df["lat_decimal_calculada"].notna() &
        (df["validacion_b2"] != "Error")
    )
    registros = df[mask].index
    print(f"  Consultando elevación para {len(registros)} registros...")

    for i, idx in enumerate(registros):
        row = df.loc[idx]
        lat = row["lat_decimal_calculada"]
        lon = row["lon_decimal_calculada"]

        elev_api, fuente = obtener_elevacion_api(lat, lon)

        df.at[idx, "elevacion_api"]    = elev_api
        df.at[idx, "elevacion_fuente"] = fuente

        if elev_api is None:
            df.at[idx, "elevacion_estado"] = "Revisar"
            df.at[idx, "elevacion_nota"]   = fuente
        else:
            # Comparar con elevación original si existe
            elev_orig = row.get(col_elev_min)
            if pd.notna(elev_orig):
                try:
                    diferencia = abs(float(elev_orig) - float(elev_api))
                    if diferencia <= 100:
                        df.at[idx, "elevacion_estado"] = "OK"
                        df.at[idx, "elevacion_nota"]   = \
                            f"Diferencia {diferencia:.0f}m con elevación original — dentro del umbral"
                    else:
                        df.at[idx, "elevacion_estado"] = "Revisar"
                        df.at[idx, "elevacion_nota"]   = \
                            f"Diferencia {diferencia:.0f}m con elevación original — supera 100m, curador decide"
                except (ValueError, TypeError):
                    df.at[idx, "elevacion_estado"] = "OK"
                    df.at[idx, "elevacion_nota"]   = f"API: {elev_api}m (sin elevación original para comparar)"
            else:
                df.at[idx, "elevacion_estado"] = "OK"
                df.at[idx, "elevacion_nota"]   = f"API: {elev_api}m (sin elevación original para comparar)"

        # Respetar límite de la API
        time.sleep(0.5)
        if (i + 1) % 20 == 0:
            print(f"    {i+1}/{len(registros)} procesados...")

    print("\n  Resultados Bloque 3:")
    print(df["elevacion_estado"].value_counts().to_string())
    return df
"""
BLOQUE 5 — Campos obligatorios Darwin Core
Protocolo SiB Colombia / Instituto Humboldt — Tabla 6 y Tabla 12 del manual
© 2026 Ximena Bedoya Araque · Universidad CES · Colecciones Biológicas CBUCES

Regla: si base en español → crear campos en español
       si base en DwC inglés → crear campos en inglés estándar Darwin Core
"""

import pandas as pd


def detectar_idioma(df):
    cols = [c.lower() for c in df.columns]
    return "dwc" if ("locality" in cols or "decimallatitude" in cols) else "es"


# Campos obligatorios según Tabla 6 y Tabla 12 del manual
# Formato: (nombre_español, nombre_dwc, descripcion)
CAMPOS_OBLIGATORIOS = [
    ("Datum",                              "geodeticDatum",
     "Debe ser WGS 84 en todos los registros"),
    ("Incertidumbre de coordenadas (m)",   "coordinateUncertaintyInMeters",
     "Toda localidad verificada o georreferenciada — excepto Nivel 7"),
    ("Comentarios de la georreferenciación","georeferenceRemarks",
     "Toda localidad sin excepción debe tener comentarios"),
    ("Fuentes de georreferenciación",      "georeferenceSources",
     "Mapa, DEM, plancha IGAC, etc."),
    ("Protocolo de georreferenciación",    "georeferenceProtocol",
     "Citar la metodología usada"),
    ("Georreferenciado por",               "georeferencedBy",
     "Nombre del georreferenciador"),
    ("Fecha de georreferenciación",        "georeferenceDate",
     "Fecha en formato ISO 8601: YYYY-MM-DD"),
    ("Plancha IGAC",                       "Plancha IGAC",
     "Número de plancha — coincide con fuentes de georreferenciación"),
]

# Mapeo español ↔ DwC para campos que YA existen con nombre diferente
MAPEO_ES_DWC = {
    # español → dwc
    "*Localidad estandarizada":       "locality",
    "Latitud decimal":                "decimalLatitude",
    "Longitud decimal":               "decimalLongitude",
    "Datum":                          "geodeticDatum",
    "Precisión (m)":                  "coordinateUncertaintyInMeters",
    "Elevación mínima (msnm)":        "minimumElevationInMeters",
    "Elevación máxima (msnm)":        "maximumElevationInMeters",
    "Latitud original":               "verbatimLatitude",
    "Longitud original":              "verbatimLongitude",
    "*País":                          "country",
    "*Departamento":                  "stateProvince",
    "*Municipio":                     "county",
    "Recurso de georreferenciación":  "georeferenceSources",
    "Comentarios de la ubicación":    "georeferenceRemarks",
}


def aplicar_bloque5(df, idioma=None):
    """
    Verifica campos obligatorios y crea los que faltan vacíos.
    Regla: nombres en español si base es española, en DwC si base es inglés.
    NO sobreescribe 'Precisión (m)' — crea columna nueva de incertidumbre.
    """
    if idioma is None:
        idioma = detectar_idioma(df)

    df = df.copy()
    reporte = []

    print(f"  Idioma detectado: {idioma}")

    for nombre_es, nombre_dwc, descripcion in CAMPOS_OBLIGATORIOS:
        # Elegir nombre según idioma
        nombre_campo = nombre_es if idioma == "es" else nombre_dwc

        # Caso especial: 'Incertidumbre de coordenadas (m)' no debe confundirse con 'Precisión (m)'
        if nombre_campo == "Incertidumbre de coordenadas (m)":
            if nombre_campo not in df.columns:
                df[nombre_campo] = ""
                reporte.append({
                    "campo": nombre_campo,
                    "estado": "creado vacío",
                    "vacios": len(df),
                    "nota": "NUEVO — distinto de 'Precisión (m)' — llenar con Bloque 4"
                })
            else:
                vacios = df[nombre_campo].isna().sum() + (df[nombre_campo].astype(str).str.strip()=="").sum()
                reporte.append({"campo": nombre_campo, "estado": "existe", "vacios": vacios, "nota": ""})
            continue

        # Para los demás campos
        if nombre_campo not in df.columns:
            # Buscar si existe con otro nombre (español ↔ DwC)
            nombre_alt = nombre_dwc if idioma == "es" else nombre_es
            if nombre_alt in df.columns:
                # Renombrar al nombre correcto según idioma
                df = df.rename(columns={nombre_alt: nombre_campo})
                vacios = df[nombre_campo].isna().sum() + (df[nombre_campo].astype(str).str.strip()=="").sum()
                reporte.append({
                    "campo": nombre_campo,
                    "estado": f"renombrado desde '{nombre_alt}'",
                    "vacios": vacios,
                    "nota": descripcion
                })
            else:
                df[nombre_campo] = ""
                reporte.append({
                    "campo": nombre_campo,
                    "estado": "creado vacío",
                    "vacios": len(df),
                    "nota": descripcion
                })
        else:
            vacios = df[nombre_campo].isna().sum() + (df[nombre_campo].astype(str).str.strip()=="").sum()
            reporte.append({
                "campo": nombre_campo,
                "estado": "existe",
                "vacios": vacios,
                "nota": descripcion
            })

    # Imprimir reporte
    print(f"\n  Estado campos obligatorios (Tabla 6 y 12 del manual):")
    for r in reporte:
        icono = "✅" if r["vacios"] == 0 else ("⚠" if r["estado"] == "existe" else "→")
        print(f"    {icono} {r['campo']}: {r['estado']} | vacíos: {r['vacios']}")

    return df, reporte


"""
BLOQUE 6 — Verificación verbatimLocality vs locality (3 casos del manual)
Protocolo SiB Colombia / Instituto Humboldt — Sección 3.6
© 2026 Ximena Bedoya Araque · Universidad CES · Colecciones Biológicas CBUCES

Caso 1: Sin localidad original + estandarizada = "Sin datos" → ✅ correcto
Caso 2: Sin localidad original + sin estandarizada            → ❌ no se estandarizó
Caso 3: Sin localidad original + estandarizada con info       → ⚠ inconsistente
Normal: Con localidad original (cualquier estandarizada)      → ✅ OK
"""

import pandas as pd


def verificar_localidad(verbatim, locality):
    """
    Retorna (estado, caso, nota)
    """
    sin_verbatim = pd.isna(verbatim) or str(verbatim).strip() in ("", "nan")
    sin_locality = pd.isna(locality) or str(locality).strip() in ("", "nan")
    loc_str = str(locality).strip().lower() if not sin_locality else ""

    # Tiene localidad original → correcto
    if not sin_verbatim:
        return "OK", "normal", "Localidad original presente"

    # Sin localidad original — aplicar 3 casos
    if sin_locality:
        return "Error", "Caso 2",
        "Sin localidad original y sin localidad estandarizada — no se realizó estandarización"

    if loc_str in ("sin datos", "sin dato", "sd"):
        return "OK", "Caso 1",
        "Sin localidad original — estandarizada correctamente como 'Sin datos'"

    return "Revisar", "Caso 3", \
        f"Sin localidad original pero estandarizada tiene información: '{str(locality)[:60]}' — verificar"


def aplicar_bloque6(df, idioma="es"):
    if idioma == "es":
        col_verbatim = "*Localidad"
        col_locality = "*Localidad estandarizada"
    else:
        col_verbatim = "verbatimLocality"
        col_locality = "locality"

    df = df.copy()
    estados, casos, notas = [], [], []

    for _, row in df.iterrows():
        verbatim = row.get(col_verbatim, "")
        locality  = row.get(col_locality, "")
        est, caso, nota = verificar_localidad(verbatim, locality)
        estados.append(est)
        casos.append(caso)
        notas.append(nota)

    df["B6_estado_localidad"] = estados
    df["B6_caso"]             = casos
    df["B6_nota_localidad"]   = notas

    print("\n  Resultados Bloque 6:")
    print(df["B6_estado_localidad"].value_counts().to_string())
    print()
    print(df["B6_caso"].value_counts().to_string())
    return df
"""
BLOQUE 7 — Clasificación inicial de niveles de calidad
Protocolo SiB Colombia / Instituto Humboldt — Tabla 9 del manual
© 2026 [Tu nombre] · Universidad CES · Colecciones Biológicas CBUCES
"""

import pandas as pd
import re
import unicodedata

# ─────────────────────────────────────────────
# Utilidades
# ─────────────────────────────────────────────

def normalizar(texto):
    """Quita tildes y convierte a minúsculas para comparar."""
    if pd.isna(texto):
        return ""
    texto = str(texto).strip()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return texto.lower()


# ─────────────────────────────────────────────
# Patrones de clasificación (Tabla 9 del manual)
# ─────────────────────────────────────────────

# NIVEL 2 — distancia u orientación explícita
# Captura todos los estilos reales de investigadores colombianos:
# "km 10", "Km. 12", "2 km", "3,5 km", "10 kilómetros", "500 m",
# "al NE", "al norte de", "1 km al SW de Nemocón", "aprox. 3 km", etc.
PATRON_NIVEL2 = re.compile(
    r"""
    \bkm\.?\s*\d+                       # km. 10 / km10 / km 10
    | \d+[.,]?\d*\s*km\.?\b             # 2 km / 3.5 km / 2,5 km
    | \bkm\.?\s+\d                      # Km. 45 via...
    | \d+[.,]?\d*\s*kil[oó]metros?\b    # 10 kilómetros
    | \d+[.,]?\d*\s*m\b(?!\w)           # 500 m (no "metros" como prefijo)
    | \d+[.,]?\d*\s*metros?\b           # 300 metros
    | \b(aprox\.?|cerca\s*de|alrededor\s*de)\s*\d   # aprox. 3 km
    | \bal\s+(N|S|E|W|NE|NW|SE|SW|NNE|NNW|SSE|SSW|ENE|ESE|WNW|WSW)\b
    | \b(norte|sur|este|oeste|noreste|noroeste|sureste|suroeste|
         nor[\s\-]?este|nor[\s\-]?oeste|sur[\s\-]?este|sur[\s\-]?oeste)\b
    | \b(al\s+)?(norte|sur|este|oeste)\s+(de|del)\b
    """,
    re.VERBOSE | re.IGNORECASE
)

# NIVEL 3 — entidad geográfica específica sin distancia
# Palabras clave que identifican un lugar concreto a nivel sub-municipal
PALABRAS_NIVEL3 = [
    "finca", "hacienda", "vereda", "corregimiento", "barrio",
    "cerro", "cano", "quebrada", "rio ", "laguna", "cienaga",
    "sector", "parcelacion", "urbanizacion", "conjunto",
    "refugio", "nucleo forestal", "reserva",
    "puente", "caserio", "inspeccion",
    "via entre", "variante", "carretera", "autopista",
    "ecoparque", "parque municipal",
    "mercarey", "cvz", "campus", "sede",  # entidades conocidas CES
]

# NIVEL 4 — áreas de jurisdicción amplia (varios municipios) o Sin datos+municipio
PALABRAS_NIVEL4 = [
    "parque nacional", "parque natural", "pnn ",
    "parque regional",                          # abarca varios municipios
    "area natural unica", "area protegida",
    "paramo", "humedal", "resguardo", "sabana", "llanura",
    "reserva forestal", "distrito de manejo",
    "cuenca", "subcuenca",
]

# NIVEL 5 — solo departamento o entidades nacionales
PALABRAS_NIVEL5 = [
    "llanos orientales", "amazonia", "orinoquia", "caribe",
    "andina", "pacifico", "region andina",
    "rio magdalena", "rio cauca", "rio meta",
    "via panamericana", "troncal",
]

# NIVEL 7 — texto explícito de información insuficiente
PATRON_NIVEL7 = re.compile(
    r"nivel\s*7|sin\s*datos.*nivel\s*7|informaci[oó]n\s*insuficiente|dudoso|ambiguo",
    re.IGNORECASE
)


# ─────────────────────────────────────────────
# Función principal de clasificación
# ─────────────────────────────────────────────

def clasificar_nivel(row, col_localidad, col_lat, col_municipio,
                     col_departamento, col_pais, col_altura):
    """
    Clasifica un registro en niveles 1–7 según Tabla 9 del manual.
    Esta es la clasificación INICIAL (Bloque 7).
    El Bloque 8 puede reclasificar Nivel 1 según validación espacial.

    Retorna: (nivel_int, justificacion_str)
    """
    loc_raw   = row.get(col_localidad, "")
    lat       = row.get(col_lat, None)
    municipio = row.get(col_municipio, "")
    depto     = row.get(col_departamento, "")
    pais      = row.get(col_pais, "")
    altura    = row.get(col_altura, None)

    loc_norm  = normalizar(loc_raw)
    muni_norm = normalizar(municipio)
    dept_norm = normalizar(depto)

    tiene_coord    = pd.notna(lat) and str(lat).strip() not in ("", "nan")
    tiene_municipio = muni_norm not in ("", "nan")
    tiene_depto    = dept_norm not in ("", "nan")
    tiene_pais     = normalizar(pais) not in ("", "nan")
    tiene_altura   = pd.notna(altura) and str(altura).strip() not in ("", "nan")
    es_sin_datos   = loc_norm in ("sin datos", "sin dato", "sd", "")

    # ── NIVEL 7: texto explícito de imposible georreferenciación ──────────
    if PATRON_NIVEL7.search(str(loc_raw)):
        return 7, "Localidad marcada explícitamente como nivel 7 o información dudosa"

    # Sin municipio, sin departamento, sin país → nivel 7
    if not tiene_municipio and not tiene_depto and not tiene_pais:
        return 7, "Sin referencia geográfica en ningún campo"

    # ── NIVEL 1: tiene coordenadas originales ─────────────────────────────
    if tiene_coord:
        return 1, "Tiene coordenadas originales — verificación espacial pendiente (Bloque 8)"

    # A partir de aquí: sin coordenadas
    # ── NIVEL 2: distancia u orientación explícita ────────────────────────
    if not es_sin_datos and PATRON_NIVEL2.search(loc_norm):
        return 2, f"Localidad con distancia u orientación: '{loc_raw}'"

    # ── NIVEL 4: área de jurisdicción amplia (EVALUAR ANTES que Nivel 3) ──
    if not es_sin_datos:
        for palabra in PALABRAS_NIVEL4:
            if palabra in loc_norm:
                return 4, f"Área de jurisdicción amplia ('{palabra}'): '{loc_raw}'"

    # ── NIVEL 3: entidad geográfica sub-municipal sin distancia ───────────
    if not es_sin_datos:
        for palabra in PALABRAS_NIVEL3:
            if palabra in loc_norm:
                return 3, f"Entidad geográfica local ('{palabra}'): '{loc_raw}'"
        # Si tiene texto pero no matchea ninguna palabra clave,
        # aún puede ser Nivel 3 (nombre propio de lugar)
        if len(loc_norm) > 3 and tiene_municipio:
            return 3, f"Nombre de lugar específico sin distancia: '{loc_raw}'"

    if es_sin_datos and tiene_municipio:
        if tiene_altura:
            # Sin datos + municipio + altura → sigue siendo Nivel 3
            # según Tabla 9: "Sin datos reportan municipio y altura"
            return 3, "Sin datos de localidad pero reporta municipio y altura — incertidumbre reducible"
        else:
            return 4, "Sin datos de localidad con municipio pero sin altura"

    # ── NIVEL 5: solo departamento o entidades nacionales ─────────────────
    for palabra in PALABRAS_NIVEL5:
        if palabra in loc_norm:
            return 5, f"Entidad geográfica a escala nacional: '{loc_raw}'"

    if not tiene_municipio and tiene_depto:
        return 5, "Sin municipio pero con departamento"

    # ── NIVEL 6: solo país ────────────────────────────────────────────────
    if not tiene_municipio and not tiene_depto and tiene_pais:
        return 6, "Solo país como referencia"

    # Fallback: si llegó hasta acá con municipio pero localidad ambigua
    return 4, f"Localidad ambigua — se asigna nivel 4 por defecto: '{loc_raw}'"


# ─────────────────────────────────────────────
# Aplicar a un DataFrame
# ─────────────────────────────────────────────

def aplicar_bloque7(df, idioma="es"):
    """
    Aplica la clasificación de niveles a todo el DataFrame.
    idioma: 'es' (español) o 'dwc' (Darwin Core inglés)
    """
    if idioma == "es":
        col_loc    = "*Localidad estandarizada"
        col_lat    = "Latitud original"
        col_muni   = "*Municipio"
        col_depto  = "*Departamento"
        col_pais   = "*País"
        col_alt    = "Altura (m)"
    else:
        col_loc    = "locality"
        col_lat    = "verbatimLatitude"
        col_muni   = "county"
        col_depto  = "stateProvince"
        col_pais   = "country"
        col_alt    = "verbatimElevation"

    resultados = df.apply(
        lambda row: clasificar_nivel(
            row, col_loc, col_lat, col_muni, col_depto, col_pais, col_alt
        ),
        axis=1,
        result_type="expand"
    )

    df = df.copy()
    df["Nivel_inicial"]              = resultados[0].astype(int)
    df["Justificacion_nivel"]        = resultados[1]

    return df


# ─────────────────────────────────────────────
# PRUEBA con datos reales
# ─────────────────────────────────────────────

"""
BLOQUE 8 — Reclasificación después de verificar coordenadas
Protocolo SiB Colombia / Instituto Humboldt
© 2026 Ximena Bedoya Araque · Universidad CES · Colecciones Biológicas CBUCES
"""

import pandas as pd
import geopandas as gpd
import numpy as np
import unicodedata
import re
import os
from shapely.geometry import Point


def normalizar(texto):
    if pd.isna(texto): return ""
    t = unicodedata.normalize("NFD", str(texto).strip())
    return "".join(c for c in t if unicodedata.category(c) != "Mn").lower()


def _insertar_punto(valor_str, es_lon=False):
    """Inserta el punto decimal en la posición correcta según rango Colombia."""
    s = str(valor_str).strip().replace(" ", "")
    negativo = s.startswith("-")
    s_abs = s.lstrip("-")
    for pos in range(1, min(len(s_abs), 4)):
        try:
            candidato = float(s_abs[:pos] + "." + s_abs[pos:])
            if negativo: candidato = -candidato
            if es_lon:
                if -82 <= candidato <= -66 or 66 <= candidato <= 82:
                    return round(-abs(candidato), 6)
            else:
                if -5 <= candidato <= 14:
                    return round(candidato, 6)
        except Exception:
            pass
    return None


def convertir_gms(texto):
    """GMS → decimal. Acepta: 06°04'47.3'' / 7°51.158'N"""
    texto = str(texto).strip()
    negativo = any(c in texto.upper() for c in ["S", "W", "O"])
    nums = re.findall(r"[\d]+[.,]?[\d]*", texto)
    if not nums: return None
    try:
        if len(nums) >= 3:
            d = float(nums[0]) + float(nums[1].replace(",",".")) / 60 + float(nums[2].replace(",",".")) / 3600
        elif len(nums) == 2:
            d = float(nums[0]) + float(nums[1].replace(",",".")) / 60
        else:
            d = float(nums[0].replace(",","."))
        return -d if negativo else d
    except Exception:
        return None


def obtener_coordenadas_decimales(row, col_lat_o, col_lon_o, col_lat_d, col_lon_d):
    """
    Convierte coordenadas a WGS84 decimal.
    Retorna: (lat, lon, fuente)
    """
    lat_o = str(row.get(col_lat_o, "")).strip()
    lon_o = str(row.get(col_lon_o, "")).strip()
    if lat_o in ("", "nan") or lon_o in ("", "nan"):
        return None, None, "sin coordenadas"

    # 1. Enteros sin punto — formato más común en la base
    if re.match(r"^-?\d{5,10}$", lat_o):
        lat_d = _insertar_punto(lat_o, es_lon=False)
        lon_d = _insertar_punto(lon_o, es_lon=True)
        if lat_d and lon_d:
            return lat_d, lon_d, "Decimal sin punto → corregido"
        return None, None, "No se pudo inferir punto decimal"

    # 2. GMS
    if "''" in lat_o or ("°" in lat_o and "''" in lat_o):
        lat_d = convertir_gms(lat_o)
        lon_d = convertir_gms(lon_o)
        if lat_d and lon_d:
            if lon_d > 0: lon_d = -lon_d
            return round(lat_d, 6), round(lon_d, 6), "GMS → decimal"
        return None, None, "Error GMS"

    # 3. GMD
    if "°" in lat_o:
        lat_d = convertir_gms(lat_o)
        lon_d = convertir_gms(lon_o)
        if lat_d and lon_d:
            if lon_d > 0: lon_d = -lon_d
            return round(lat_d, 6), round(lon_d, 6), "GMD → decimal"
        return None, None, "Error GMD"

    # 4. Decimal directo
    try:
        lat_d = float(lat_o.replace(",", "."))
        lon_d = float(lon_o.replace(",", "."))
        if lon_d > 0 and 66 <= lon_d <= 82: lon_d = -lon_d
        if -5 <= lat_d <= 14 and -82 <= lon_d <= -66:
            return round(lat_d, 6), round(lon_d, 6), "Decimal directo"
    except Exception:
        pass

    return None, None, "Formato no reconocido"


def cargar_gadm(ruta_gadm):
    gdf = gpd.read_file(ruta_gadm)
    gdf["depto_norm"] = gdf["NAME_1"].apply(normalizar)
    gdf["muni_norm"]  = gdf["NAME_2"].apply(normalizar)
    if gdf.crs is None or gdf.crs.to_epsg() != 4326:
        gdf = gdf.set_crs("EPSG:4326")
    return gdf


def validar_punto_municipio(lat, lon, municipio_rep, depto_rep, gdf_gadm):
    if lat is None or lon is None:
        return "🔵", "", "", "Sin coordenadas convertibles"
    if not (-5 <= lat <= 14 and -82 <= lon <= -66):
        if -5 <= lat <= 16 and -84 <= lon <= -60:
            return "⚠", "", "", "Zona marítima — verificar"
        return "❌", "", "", f"Fuera de Colombia: lat={lat}, lon={lon}"

    punto = Point(lon, lat)
    hits  = gdf_gadm[gdf_gadm.geometry.contains(punto)]
    if hits.empty:
        hits = gdf_gadm[gdf_gadm.geometry.distance(punto) < 0.01]
    if hits.empty:
        return "❌", "", "", "Punto fuera de todos los municipios"

    muni_det  = hits.iloc[0]["NAME_2"]
    depto_det = hits.iloc[0]["NAME_1"]

    if normalizar(municipio_rep) == normalizar(muni_det):
        return "✅", muni_det, depto_det, f"Dentro de {muni_det}, {depto_det}"
    if normalizar(depto_rep) == normalizar(depto_det):
        return "⚠", muni_det, depto_det, \
            f"Reportado: {municipio_rep} | Detectado: {muni_det} ({depto_det})"
    return "❌", muni_det, depto_det, \
        f"Reportado: {municipio_rep} ({depto_rep}) | Detectado: {muni_det} ({depto_det})"


def aplicar_bloque8(df, ruta_gadm, idioma="es"):
    if idioma == "es":
        col_lat_o = "Latitud original";  col_lon_o = "Longitud original"
        col_lat_d = "Latitud decimal";   col_lon_d = "Longitud decimal"
        col_muni  = "*Municipio";        col_depto = "*Departamento"
    else:
        col_lat_o = "verbatimLatitude";  col_lon_o = "verbatimLongitude"
        col_lat_d = "decimalLatitude";   col_lon_d = "decimalLongitude"
        col_muni  = "county";            col_depto = "stateProvince"

    print("  Cargando capa GADM Colombia...")
    gdf_gadm = cargar_gadm(ruta_gadm)
    print(f"  {len(gdf_gadm)} municipios cargados")

    df = df.copy()
    df["lat_wgs84"]           = np.nan
    df["lon_wgs84"]           = np.nan
    df["fuente_conversion"]   = ""
    df["validacion_b2"]       = ""
    df["municipio_detectado"] = ""
    df["depto_detectado"]     = ""
    df["mensaje_b2"]          = ""
    df["Nivel_final"]         = df["Nivel_inicial"]

    nivel1_idx = df[df["Nivel_inicial"] == 1].index
    print(f"  Verificando {len(nivel1_idx)} registros de Nivel 1...")

    for idx in nivel1_idx:
        row = df.loc[idx]
        lat, lon, fuente = obtener_coordenadas_decimales(
            row, col_lat_o, col_lon_o, col_lat_d, col_lon_d)

        df.at[idx, "lat_wgs84"]         = lat
        df.at[idx, "lon_wgs84"]         = lon
        df.at[idx, "fuente_conversion"] = fuente

        estado, muni_d, depto_d, mensaje = validar_punto_municipio(
            lat, lon, row.get(col_muni, ""), row.get(col_depto, ""), gdf_gadm)

        df.at[idx, "validacion_b2"]       = estado
        df.at[idx, "municipio_detectado"] = muni_d
        df.at[idx, "depto_detectado"]     = depto_d
        df.at[idx, "mensaje_b2"]          = mensaje

        if estado == "❌":
            df.at[idx, "Nivel_final"] = 4
            df.at[idx, "Justificacion_nivel"] = f"Reclasificado: {mensaje}"
        else:
            df.at[idx, "Nivel_final"] = 1

    print("  Bloque 8 completado.")
    return df


"""
BLOQUE 9 — Asignación de coordenadas para registros sin coordenadas (Niveles 2-6)
Protocolo SiB Colombia / Instituto Humboldt — Tabla 9 y sección 3.5.2 del manual
© 2026 Ximena Bedoya Araque · Universidad CES · Colecciones Biológicas CBUCES

Estrategia por nivel (respaldada por manual sección 3.5.2 y Tabla 9):
  Nivel 2 → Nominatim (base de datos de topónimos OSM) → fallback centroide municipio
  Nivel 3 → Centroide del municipio (GADM)
  Nivel 4 → Centroide del municipio (GADM)
  Nivel 5 → Centroide del departamento (GADM)
  Nivel 6 → Centroide de Colombia (fijo: 4.5°N, -74.2°W)
  Nivel 7 → Sin coordenadas (no se georreferencia — manual Tabla 9)

Fuente documentada en campo "Fuentes de georreferenciación" / georeferenceSources:
  Nominatim → "OpenStreetMap Nominatim (gacetero de topónimos)"
  GADM      → "Centroide municipio GADM v4.1 (proxy conservador)"
  Colombia  → "Centroide Colombia — alta incertidumbre"
"""

import pandas as pd
import geopandas as gpd
import numpy as np
import unicodedata
import re
import time
import requests
from shapely.geometry import Point


# ─────────────────────────────────────────────────────────────────
# Utilidades
# ─────────────────────────────────────────────────────────────────

def normalizar(texto):
    if pd.isna(texto): return ""
    t = unicodedata.normalize("NFD", str(texto).strip())
    return "".join(c for c in t if unicodedata.category(c) != "Mn").lower()


def detectar_idioma(df):
    """
    Detecta si la base está en español o Darwin Core inglés.
    Retorna: 'es' o 'dwc'
    """
    cols = [c.lower() for c in df.columns]
    if "locality" in cols or "decimallatitude" in cols:
        return "dwc"
    return "es"


def nombres_columnas(idioma):
    """
    Retorna diccionario de nombres de columnas según idioma.
    Regla: si base en español → nombres en español.
           si base en DwC inglés → nombres en inglés estándar Darwin Core.
    """
    if idioma == "es":
        return {
            "lat_final":        "Latitud georreferenciada",
            "lon_final":        "Longitud georreferenciada",
            "fuentes":          "Fuentes de georreferenciación",
            "comentarios":      "Comentarios de la ubicación",
            "protocolo":        "Protocolo de georreferenciación",
            "por":              "Georreferenciado por",
            "fecha":            "Fecha de georreferenciación",
            "incertidumbre":    "Incertidumbre de coordenadas (m)",
            "municipio":        "*Municipio",
            "departamento":     "*Departamento",
            "pais":             "*País",
            "localidad":        "*Localidad estandarizada",
            "altura":           "Altura (m)",
        }
    else:
        return {
            "lat_final":        "decimalLatitude",
            "lon_final":        "decimalLongitude",
            "fuentes":          "georeferenceSources",
            "comentarios":      "georeferenceRemarks",
            "protocolo":        "georeferenceProtocol",
            "por":              "georeferencedBy",
            "fecha":            "georeferenceDate",
            "incertidumbre":    "coordinateUncertaintyInMeters",
            "municipio":        "county",
            "departamento":     "stateProvince",
            "pais":             "country",
            "localidad":        "locality",
            "altura":           "verbatimElevation",
        }


# ─────────────────────────────────────────────────────────────────
# Geocodificación con Nominatim (Nivel 2)
# ─────────────────────────────────────────────────────────────────

NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
HEADERS = {"User-Agent": "verificador-georef-sib/1.0 (CES Colombia, educativo)"}

def geocodificar_nominatim(localidad, municipio, departamento, pais="Colombia"):
    """
    Intenta geocodificar una localidad usando Nominatim (OpenStreetMap).
    Estrategia:
      1. Localidad completa + municipio + departamento + Colombia
      2. Solo municipio + departamento + Colombia (fallback)
    Retorna: (lat, lon, fuente, precision) o (None, None, msg, None)
    """
    intentos = [
        f"{localidad}, {municipio}, {departamento}, {pais}",
        f"{municipio}, {departamento}, {pais}",
    ]
    etiquetas = [
        "Nominatim — localidad completa",
        "Nominatim — centroide municipio (fallback)"
    ]

    for query, etiqueta in zip(intentos, etiquetas):
        try:
            params = {
                "q": query,
                "format": "json",
                "limit": 1,
                "countrycodes": "co",
                "addressdetails": 0,
            }
            r = requests.get(NOMINATIM_URL, params=params,
                             headers=HEADERS, timeout=8)
            if r.status_code == 200:
                data = r.json()
                if data:
                    lat = float(data[0]["lat"])
                    lon = float(data[0]["lon"])
                    # Validar que caiga en Colombia
                    if -5 <= lat <= 16 and -82 <= lon <= -60:
                        return lat, lon, f"OpenStreetMap Nominatim — {etiqueta}", data[0].get("type","")
            time.sleep(1)  # respetar límite Nominatim: 1 req/segundo
        except Exception:
            pass

    return None, None, "Nominatim no disponible", None


# ─────────────────────────────────────────────────────────────────
# Centroides desde GADM
# ─────────────────────────────────────────────────────────────────

def calcular_centroides_gadm(gdf_gadm):
    """
    Pre-calcula centroides de municipios y departamentos desde GADM.
    Retorna dos diccionarios:
      centroides_muni[nombre_norm]  → (lat, lon)
      centroides_depto[nombre_norm] → (lat, lon)
    """
    # Proyectar a EPSG:3116 para cálculo correcto, luego volver a WGS84
    gdf_proj = gdf_gadm.to_crs("EPSG:3116")

    centroides_muni = {}
    for _, row in gdf_proj.iterrows():
        c = row.geometry.centroid
        # Convertir de vuelta a WGS84
        punto_wgs = gpd.GeoSeries([c], crs="EPSG:3116").to_crs("EPSG:4326").iloc[0]
        key = normalizar(row["NAME_2"])
        centroides_muni[key] = (round(punto_wgs.y, 6), round(punto_wgs.x, 6))

    # Centroides de departamentos (disolver por NAME_1)
    gdf_deptos = gdf_proj.dissolve(by="NAME_1").reset_index()
    centroides_depto = {}
    for _, row in gdf_deptos.iterrows():
        c = row.geometry.centroid
        punto_wgs = gpd.GeoSeries([c], crs="EPSG:3116").to_crs("EPSG:4326").iloc[0]
        key = normalizar(row["NAME_1"])
        centroides_depto[key] = (round(punto_wgs.y, 6), round(punto_wgs.x, 6))

    return centroides_muni, centroides_depto


# ─────────────────────────────────────────────────────────────────
# Función principal Bloque 9
# ─────────────────────────────────────────────────────────────────

def aplicar_bloque9(df, ruta_gadm, idioma=None, usar_nominatim=True):
    """
    Asigna coordenadas a registros sin coordenadas (Niveles 2-6).

    Parámetros:
      df           : DataFrame con columna Nivel_final del Bloque 8
      ruta_gadm    : ruta al gadm41_COL_2.json
      idioma       : 'es' o 'dwc' — si None, se detecta automáticamente
      usar_nominatim: True en producción (Streamlit Cloud),
                      False en pruebas sin red
    """
    if idioma is None:
        idioma = detectar_idioma(df)

    cols = nombres_columnas(idioma)
    print(f"  Idioma detectado: {idioma}")

    # Cargar GADM y calcular centroides
    print("  Cargando GADM y calculando centroides...")
    gdf_gadm = gpd.read_file(ruta_gadm)
    if gdf_gadm.crs is None or gdf_gadm.crs.to_epsg() != 4326:
        gdf_gadm = gdf_gadm.set_crs("EPSG:4326")
    centroides_muni, centroides_depto = calcular_centroides_gadm(gdf_gadm)
    print(f"  Centroides: {len(centroides_muni)} municipios, {len(centroides_depto)} departamentos")

    df = df.copy()

    # Inicializar columnas de resultado si no existen
    for col_key in ["lat_final", "lon_final", "fuentes", "comentarios"]:
        col_name = cols[col_key]
        if col_name not in df.columns:
            df[col_name] = ""

    # Procesar solo registros sin coordenadas convertidas (lat_wgs84 vacío)
    sin_coord_mask = (
        df["Nivel_final"].isin([2, 3, 4, 5, 6]) &
        (df["lat_wgs84"].isna() | (df["lat_wgs84"].astype(str) == "nan"))
    )

    registros_procesar = df[sin_coord_mask].index
    print(f"  Registros a georreferenciar: {len(registros_procesar)}")
    print(f"    Nivel 2: {(df.loc[registros_procesar,'Nivel_final']==2).sum()}")
    print(f"    Nivel 3: {(df.loc[registros_procesar,'Nivel_final']==3).sum()}")
    print(f"    Nivel 4: {(df.loc[registros_procesar,'Nivel_final']==4).sum()}")
    print(f"    Nivel 5: {(df.loc[registros_procesar,'Nivel_final']==5).sum()}")
    print(f"    Nivel 6: {(df.loc[registros_procesar,'Nivel_final']==6).sum()}")

    ok = 0
    fallback = 0
    sin_resultado = 0

    for idx in registros_procesar:
        row    = df.loc[idx]
        nivel  = int(row["Nivel_final"])
        muni   = str(row.get(cols["municipio"], "")).strip()
        depto  = str(row.get(cols["departamento"], "")).strip()
        # Buscar columna de país con y sin tilde
        pais   = str(row.get("*País", row.get("*Pais", row.get("country", "Colombia")))).strip()
        loc    = str(row.get(cols["localidad"], "")).strip()

        lat, lon, fuente, comentario = None, None, "", ""

        # ── NIVEL 2: Nominatim primero ─────────────────────────────
        if nivel == 2:
            if usar_nominatim:
                lat, lon, fuente, tipo = geocodificar_nominatim(loc, muni, depto, pais)
                if lat:
                    comentario = (
                        f"Nivel 2 — coordenada inferida desde descripción de localidad "
                        f"usando base de topónimos OSM. "
                        f"Verificar coincidencia con municipio {muni}."
                    )
                    ok += 1
                else:
                    # Fallback a centroide municipio
                    lat, lon = centroides_muni.get(normalizar(muni), (None, None))
                    fuente = "Centroide municipio GADM v4.1 — Nominatim no disponible"
                    comentario = (
                        f"Nivel 2 — coordenada provisional: centroide de {muni}. "
                        f"Georreferenciación exacta requiere verificación manual."
                    )
                    fallback += 1
            else:
                # Sin red: directo a centroide
                lat, lon = centroides_muni.get(normalizar(muni), (None, None))
                fuente = "Centroide municipio GADM v4.1 (sin acceso a Nominatim)"
                comentario = f"Nivel 2 — centroide provisional de {muni}."
                fallback += 1

        # ── NIVEL 3 y 4: centroide municipio ──────────────────────
        elif nivel in [3, 4]:
            lat, lon = centroides_muni.get(normalizar(muni), (None, None))
            fuente = "Centroide municipio GADM v4.1 (proxy conservador)"
            comentario = (
                f"Nivel {nivel} — centroide de {muni}, {depto}. "
                f"Incertidumbre = radio máximo del municipio."
            )
            if lat: ok += 1
            else: sin_resultado += 1

        # ── NIVEL 5: centroide departamento ───────────────────────
        elif nivel == 5:
            lat, lon = centroides_depto.get(normalizar(depto), (None, None))
            fuente = "Centroide departamento GADM v4.1"
            comentario = (
                f"Nivel 5 — centroide de {depto}. Alta incertidumbre."
            )
            if lat: ok += 1
            else: sin_resultado += 1

        # ── NIVEL 6: centroide Colombia ────────────────────────────
        elif nivel == 6:
            lat, lon = 4.570868, -74.297333  # centroide Colombia
            fuente = "Centroide Colombia — alta incertidumbre"
            comentario = "Nivel 6 — solo país como referencia. Muy alta incertidumbre."
            ok += 1

        # Guardar resultado
        if lat is not None and lon is not None:
            df.at[idx, cols["lat_final"]]  = f"{float(lat):.6f}"
            df.at[idx, cols["lon_final"]]  = f"{float(lon):.6f}"
            df.at[idx, cols["fuentes"]]    = fuente
            df.at[idx, cols["comentarios"]] = comentario
        else:
            df.at[idx, cols["comentarios"]] = f"Sin municipio reconocido — verificar campo {cols['municipio']}"
            sin_resultado += 1

    # Para Nivel 1 confirmado: copiar lat_wgs84/lon_wgs84 a columnas finales
    nivel1_mask = df["Nivel_final"] == 1
    df.loc[nivel1_mask, cols["lat_final"]] = df.loc[nivel1_mask, "lat_wgs84"].apply(
        lambda x: f"{float(x):.6f}" if pd.notna(x) and str(x) not in ("","nan") else "")
    df.loc[nivel1_mask, cols["lon_final"]] = df.loc[nivel1_mask, "lon_wgs84"].apply(
        lambda x: f"{float(x):.6f}" if pd.notna(x) and str(x) not in ("","nan") else "")
    df.loc[nivel1_mask, cols["fuentes"]]   = df.loc[nivel1_mask, "fuente_conversion"]

    # Nivel 7: documentar que no se georreferencia
    nivel7_mask = df["Nivel_final"] == 7
    df.loc[nivel7_mask, cols["comentarios"]] = (
        "Nivel 7 — localidad con información dudosa o insuficiente. "
        "No se georreferencia según protocolo SiB Colombia."
    )

    print(f"\n  Resultados Bloque 9:")
    print(f"    Georreferenciados con fuente primaria: {ok}")
    print(f"    Con centroide provisional (fallback):  {fallback}")
    print(f"    Sin resultado (municipio no encontrado): {sin_resultado}")

    return df


# ─────────────────────────────────────────────────────────────────
# PRUEBA con datos reales
# ─────────────────────────────────────────────────────────────────

"""
BLOQUE 10 — Exportar Excel final con colores y hoja de resumen
Protocolo SiB Colombia / Instituto Humboldt — Tabla 12 y Tabla 13 del manual
© 2026 Ximena Bedoya Araque · Universidad CES · Colecciones Biológicas CBUCES

Nombre correcto según manual (Tabla 6):
  georeferenceRemarks → "Comentarios de la georreferenciación"
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import io, re, unicodedata

# ─── Colores ───────────────────────────────────────────────────
VERDE      = PatternFill("solid", fgColor="C6EFCE")
AMARILLO   = PatternFill("solid", fgColor="FFEB9C")
ROJO       = PatternFill("solid", fgColor="FFC7CE")
GRIS       = PatternFill("solid", fgColor="D9D9D9")
AZUL_DARK  = PatternFill("solid", fgColor="1F4E79")
AZUL_MED   = PatternFill("solid", fgColor="BDD7EE")
VERDE_COLS = PatternFill("solid", fgColor="E2EFDA")  # columnas nuevas del protocolo
BLANCO     = PatternFill("solid", fgColor="FFFFFF")   # sin color (blanco)

PROTOCOLO = (
    "Escobar D, Jojoa LM, Díaz SR, Rudas E, Albarracín RD, Ramírez C, "
    "Gómez JY, López CR, Saavedra J, Ortiz R (2016). "
    "Georreferenciación de localidades: Una guía de referencia para "
    "colecciones biológicas. Instituto Humboldt - ICN/UNAL. "
    "Bogotá D.C., Colombia. 144 p."
)

# Columnas nuevas que agrega el protocolo (encabezado en verde claro)
COLUMNAS_NUEVAS_ES = {
    "Nivel de calidad inicial",
    "Justificación del nivel",
    "Nivel de calidad final",
    "Resultado validación espacial",
    "Latitud georreferenciada",
    "Longitud georreferenciada",
    "Fuentes de georreferenciación",
    "Comentarios de la georreferenciación",
    "Protocolo de georreferenciación",
    "Georreferenciado por",
    "Fecha de georreferenciación",
    "Incertidumbre de coordenadas (m)",
    "Origen",
}

COLUMNAS_NUEVAS_DWC = {
    "Nivel de calidad inicial",
    "Justificación del nivel",
    "Nivel de calidad final",
    "Resultado validación espacial",
    "decimalLatitude_georef",
    "decimalLongitude_georef",
    "georeferenceSources",
    "georeferenceRemarks",
    "georeferenceProtocol",
    "georeferencedBy",
    "georeferenceDate",
    "coordinateUncertaintyInMeters",
    "Origen",
}


def detectar_idioma(df):
    cols = [c.lower() for c in df.columns]
    return "dwc" if ("locality" in cols or "decimallatitude" in cols) else "es"


def nombres_columnas(idioma):
    if idioma == "es":
        return {
            "lat_orig":      "Latitud original",
            "lon_orig":      "Longitud original",
            "lat_final":     "Latitud georreferenciada",
            "lon_final":     "Longitud georreferenciada",
            "fuentes":       "Fuentes de georreferenciación",
            "comentarios":   "Comentarios de la georreferenciación",
            "protocolo":     "Protocolo de georreferenciación",
            "por":           "Georreferenciado por",
            "fecha":         "Fecha de georreferenciación",
            "incertidumbre": "Incertidumbre de coordenadas (m)",
            "datum":         "Datum",
            "elev_min":      "Elevación mínima (msnm)",
            "elev_max":      "Elevación máxima (msnm)",
            "localidad":     "*Localidad estandarizada",
            "municipio":     "*Municipio",
            "departamento":  "*Departamento",
        }
    else:
        return {
            "lat_orig":      "verbatimLatitude",
            "lon_orig":      "verbatimLongitude",
            "lat_final":     "decimalLatitude_georef",
            "lon_final":     "decimalLongitude_georef",
            "fuentes":       "georeferenceSources",
            "comentarios":   "georeferenceRemarks",
            "protocolo":     "georeferenceProtocol",
            "por":           "georeferencedBy",
            "fecha":         "georeferenceDate",
            "incertidumbre": "coordinateUncertaintyInMeters",
            "datum":         "geodeticDatum",
            "elev_min":      "minimumElevationInMeters",
            "elev_max":      "maximumElevationInMeters",
            "localidad":     "locality",
            "municipio":     "county",
            "departamento":  "stateProvince",
        }


# ─── Corrección de coordenadas ─────────────────────────────────
def insertar_punto(valor_str, es_lon=False):
    s = str(valor_str).strip()
    negativo = s.startswith('-')
    s_abs = s.lstrip('-')
    for pos in range(1, min(len(s_abs), 4)):
        try:
            candidato = float(s_abs[:pos] + '.' + s_abs[pos:])
            if negativo: candidato = -candidato
            if es_lon:
                if -82 <= candidato <= -66 or 66 <= candidato <= 82:
                    return round(-abs(candidato), 6)
            else:
                if -5 <= candidato <= 14:
                    return round(candidato, 6)
        except: pass
    return None


def corregir_coordenada(valor, es_lon=False):
    """
    Convierte cualquier formato de coordenada a decimal WGS84.
    Maneja punto Y coma como separador decimal (según configuración del PC).
    Retorna float o None si no se puede convertir.
    """
    if pd.isna(valor): return None
    s = str(valor).strip()
    if s in ('', 'nan'): return None

    # Limpiar caracteres inválidos al inicio
    s = s.replace('?', '').replace(' ', '').strip()
    if not s: return None

    # Caso: múltiples puntos como separadores de miles
    # 6.787.898 → 6787898 o 6.787898
    # 6.913.160.858 → 6913160858
    if s.count('.') > 1:
        # Quitar negativos para analizar
        neg = s.startswith('-')
        s_abs = s.lstrip('-')
        partes = s_abs.split('.')
        # Si todas las partes internas tienen 3 dígitos → separador de miles
        internas = partes[:-1]
        ultima = partes[-1]
        if all(len(p) == 3 for p in internas[1:]):
            # El primer número puede tener 1-3 dígitos, los demás 3
            # Unir todo y agregar punto antes de los últimos dígitos
            todo = ''.join(partes)
            s = ('-' if neg else '') + todo
        else:
            # Simplemente quitar todos los puntos excepto el último
            s = ('-' if neg else '') + ''.join(partes[:-1]) + '.' + ultima

    # Coma → punto (configuración europea)
    # Pero solo si hay exactamente una coma (separador decimal)
    if s.count(',') == 1 and s.count('.') == 0:
        s = s.replace(',', '.')
    elif s.count(',') > 1:
        # Coma como separador de miles: 6,787,898 → 6787898
        s = s.replace(',', '')

    # Decimal directo válido
    try:
        f = float(s)
        if es_lon:
            if -82 <= f <= -66: return round(f, 6)
            if 66 <= f <= 82:   return round(-f, 6)
        else:
            if -5 <= f <= 14:   return round(f, 6)
    except: pass

    # Entero sin punto decimal
    s_clean = s.lstrip('-')
    if re.match(r'^\d{5,10}$', s_clean):
        return insertar_punto(s, es_lon)

    # GMS: contiene ''
    if "''" in s or "'''" in s:
        nums = re.findall(r'[\d]+[.,]?[\d]*', s)
        negativo = any(c in s.upper() for c in ['S','W','O'])
        try:
            if len(nums) >= 3:
                d = float(nums[0]) + float(nums[1].replace(',','.')) / 60 \
                    + float(nums[2].replace(',','.')) / 3600
            elif len(nums) == 2:
                d = float(nums[0]) + float(nums[1].replace(',','.')) / 60
            else:
                d = float(nums[0].replace(',','.'))
            d = -d if negativo else d
            if es_lon and d > 0: d = -d
            return round(d, 6)
        except: pass

    # GMD: contiene °
    if '°' in s:
        nums = re.findall(r'[\d]+[.,]?[\d]*', s)
        negativo = any(c in s.upper() for c in ['S','W','O'])
        try:
            if len(nums) >= 2:
                d = float(nums[0]) + float(nums[1].replace(',','.')) / 60
            else:
                d = float(nums[0].replace(',','.'))
            d = -d if negativo else d
            if es_lon and d > 0: d = -d
            return round(d, 6)
        except: pass

    return None


def preparar_coordenadas(df, idioma):
    cols = nombres_columnas(idioma)
    df = df.copy()
    col_lat_o = cols["lat_orig"]
    col_lon_o = cols["lon_orig"]
    col_lat_f = cols["lat_final"]
    col_lon_f = cols["lon_final"]

    if col_lat_f not in df.columns: df[col_lat_f] = np.nan
    if col_lon_f not in df.columns: df[col_lon_f] = np.nan

    mask = (
        df[col_lat_o].notna() &
        (df[col_lat_o].astype(str).str.strip() != '') &
        (df[col_lat_o].astype(str).str.strip() != 'nan')
    )
    df.loc[mask, col_lat_f] = pd.to_numeric(
        df.loc[mask, col_lat_o].apply(lambda x: corregir_coordenada(x, es_lon=False)),
        errors="coerce"
    )
    df.loc[mask, col_lon_f] = pd.to_numeric(
        df.loc[mask, col_lon_o].apply(lambda x: corregir_coordenada(x, es_lon=True)),
        errors="coerce"
    )

    # Si bloque 8 ya calculó lat_wgs84, usar esa (más precisa)
    if 'lat_wgs84' in df.columns:
        mask_b8 = df['lat_wgs84'].notna()
        df.loc[mask_b8, col_lat_f] = df.loc[mask_b8, 'lat_wgs84']
        df.loc[mask_b8, col_lon_f] = df.loc[mask_b8, 'lon_wgs84']
    return df


def preparar_dataframe(df, idioma):
    cols = nombres_columnas(idioma)
    df = df.copy()
    df = preparar_coordenadas(df, idioma)

    # Datum
    col_datum = cols["datum"]
    if col_datum in df.columns:
        df[col_datum] = df[col_datum].fillna("WGS 84 (asumido)")
        df.loc[df[col_datum].astype(str).str.strip() == "", col_datum] = "WGS 84 (asumido)"

    # Elevación mín = máx
    col_emin, col_emax = cols["elev_min"], cols["elev_max"]
    if col_emin in df.columns and col_emax in df.columns:
        df.loc[df[col_emin].notna() & df[col_emax].isna(), col_emax] = df[col_emin]
        df.loc[df[col_emax].notna() & df[col_emin].isna(), col_emin] = df[col_emax]

    # Protocolo en georreferenciados
    col_prot = cols["protocolo"]
    if col_prot not in df.columns: df[col_prot] = ""
    df.loc[df["Nivel_final"].isin([1,2,3,4,5,6]), col_prot] = PROTOCOLO

    # 5. Comentarios de georreferenciación — estructura del manual sección 3.6
    # "Nivel + entidad geográfica + observaciones incertidumbre + cambios geografía superior"
    col_com = cols["comentarios"]
    if col_com not in df.columns:
        df[col_com] = ""
    if "Nivel_final" not in df.columns and "Nivel_inicial" in df.columns:
        df["Nivel_final"] = df["Nivel_inicial"].copy()
    # Generar comentario específico para cada registro
    df[col_com] = df.apply(lambda row: generar_comentario(row, idioma), axis=1)
    return df


# ─── Color por fila ────────────────────────────────────────────
def color_por_fila(row):
    """
    Color basado SOLO en resultado de validacion espacial.
    Verde    -> OK (coordenada dentro del municipio)
    Amarillo -> Revisar (municipio vecino o zona maritima)
    Rojo     -> Error (fuera de Colombia)
    Gris     -> Solo Nivel 7 (no georreferenciable — manual Tabla 9)
    None     -> Sin validacion aun (niveles 2-6 sin coords originales)
    """
    nivel  = int(row.get("Nivel_final", row.get("Nivel_inicial", 0)))
    # Buscar con y sin tilde para mayor robustez
    val_b2 = str(
        row.get("Resultado validación espacial",
        row.get("Resultado validacion espacial",
        row.get("validacion_b2", "")))
    ).strip()
    if nivel == 7: return GRIS
    if val_b2 in ("", "nan"): return BLANCO
    if "Error"   in val_b2 or "❌" in val_b2: return ROJO
    if "Revisar" in val_b2 or "⚠"  in val_b2: return AMARILLO
    if "OK"      in val_b2 or "✅" in val_b2:  return VERDE
    return BLANCO


# ─── Texto de validación sin emojis (para celdas Excel) ───────
def texto_validacion(val):
    """
    Convierte resultado a texto accesible para Excel.
    Incluye indicador de forma (no solo color) para daltonismo.
    ✅ OK       → [✓] OK
    ⚠  Revisar  → [!] Revisar
    ❌ Error    → [X] Error
    """
    if pd.isna(val) or str(val).strip() == "": return ""
    s = str(val)
    s = s.replace("✅","OK").replace("⚠","Revisar").replace("❌","Error").replace("🔵","Sin coords")
    if s == "OK":      return "[✓] OK"
    if s == "Revisar": return "[!] Revisar"
    if s == "Error":   return "[X] Error"
    return s



# ─── Comentarios de georreferenciación por registro ───────────
def generar_comentario(row, idioma="es"):
    """
    Genera el comentario de georreferenciación según estructura del manual:
    "Nivel de calidad, entidad geográfica de referencia,
     observaciones sobre incertidumbre, cambios en geografía superior."
    (Manual sección 3.6, Tabla 13)
    """
    nivel     = int(row.get("Nivel_final", row.get("Nivel_inicial", 0)))
    muni      = str(row.get("*Municipio" if idioma=="es" else "county", "")).strip()
    depto     = str(row.get("*Departamento" if idioma=="es" else "stateProvince", "")).strip()
    localidad = str(row.get("*Localidad estandarizada" if idioma=="es" else "locality", "")).strip()
    formato   = str(row.get("formato_coordenada", "")).strip()
    lat       = row.get("lat_decimal_calculada", row.get("lat_wgs84", ""))
    lon       = row.get("lon_decimal_calculada", row.get("lon_wgs84", ""))
    validacion = str(row.get("validacion_b2", row.get("Resultado validación espacial", ""))).strip()
    muni_det  = str(row.get("municipio_detectado", "")).strip()
    incert    = row.get("Incertidumbre de coordenadas (m)", "")
    incert_txt = str(int(incert)) + " m" if incert and str(incert) not in ("","nan") else "pendiente de cálculo"
    just      = str(row.get("Justificacion_nivel", row.get("Justificación del nivel",""))).strip()

    lat_txt = str(round(float(lat),6)) if lat and str(lat) not in ("","nan") else "sin dato"
    lon_txt = str(round(float(lon),6)) if lon and str(lon) not in ("","nan") else "sin dato"

    if nivel == 1:
        if "Error" in validacion or "❌" in validacion:
            return (
                f"Nivel 1. Coordenada original del colector (formato: {formato}). "
                f"ERROR: coordenada ({lat_txt}, {lon_txt}) fuera del territorio colombiano. "
                f"Se reclasificó según información de localidad. Requiere corrección manual."
            )
        elif "Revisar" in validacion or "⚠" in validacion:
            return (
                f"Nivel 1. Coordenada original del colector (formato: {formato}). "
                f"Convertida a WGS84: {lat_txt}, {lon_txt}. "
                f"ADVERTENCIA: punto detectado en {muni_det} — municipio reportado: {muni}. "
                f"Posible error en límites administrativos o documentación del colector. "
                f"Verificar visualmente. Incertidumbre: {incert_txt}."
            )
        elif "maritima" in str(row.get("rango_colombia","")).lower():
            return (
                f"Nivel 1. Coordenada en zona marítima colombiana ({lat_txt}, {lon_txt}). "
                f"Verificar con fuente oficial CCO: https://cco.gov.co"
            )
        else:
            return (
                f"Nivel 1. Coordenada original del colector (formato: {formato}). "
                f"Convertida a grados decimales WGS84: {lat_txt}, {lon_txt}. "
                f"Validación espacial: punto dentro de {muni}, {depto}. "
                f"Incertidumbre calculada (Tabla 2 del manual): {incert_txt}. "
                f"Sin verificación de campo."
            )

    elif nivel == 2:
        return (
            f"Nivel 2. La localidad describe una distancia u orientación: '{localidad}'. "
            f"No tiene coordenadas originales — se buscó el punto en la base de "
            f"topónimos OpenStreetMap (Nominatim) para {muni}, {depto}. "
            f"Si Nominatim no encontró el punto exacto, se usó el centro del municipio "
            f"como referencia. Incertidumbre: {incert_txt}. "
            f"Se recomienda verificar visualmente sobre el mapa de la aplicación."
        )

    elif nivel == 3:
        return (
            f"Nivel 3. Este registro describe un lugar específico ('{localidad}') "
            f"pero no tiene coordenadas originales. "
            f"Se le asignó la coordenada del centro geográfico del municipio {muni}, {depto} "
            f"como punto de referencia (fuente: GADM v4.1). "
            f"Esta coordenada es aproximada — la incertidumbre abarca todo el municipio "
            f"({incert_txt}). Para mayor precisión, georreferenciar sobre cartografía IGAC."
        )

    elif nivel == 4:
        if localidad.lower() in ("sin datos","sin dato","sd",""):
            desc = f"No se registró localidad específica, solo municipio ({muni}, {depto})"
        else:
            desc = f"Localidad '{localidad}' corresponde a un área de amplia jurisdicción en {muni}, {depto}"
        return (
            f"Nivel 4. {desc}. "
            f"Se asignó la coordenada del centro geográfico del municipio {muni}, {depto} "
            f"(fuente: GADM v4.1). "
            f"Incertidumbre alta — abarca todo el municipio ({incert_txt}). "
            f"Se recomienda no usar para análisis espaciales de alta precisión."
        )

    elif nivel == 5:
        return (
            f"Nivel 5. Solo se conoce el departamento de colecta: {depto}. "
            f"Sin municipio ni localidad disponibles. "
            f"Se asignó la coordenada del centro geográfico del departamento {depto} "
            f"(fuente: GADM v4.1). Incertidumbre muy alta — abarca todo el departamento. "
            f"Solo útil para análisis de distribución geográfica gruesa."
        )

    elif nivel == 6:
        return (
            f"Nivel 6. Solo se sabe que el registro es de Colombia. "
            f"Sin departamento, municipio ni localidad disponibles. "
            f"Se asignó la coordenada del centro geográfico de Colombia (4.57°N, 74.30°O). "
            f"Incertidumbre máxima. Solo sirve para confirmar presencia en el país — "
            f"no usar en análisis espaciales."
        )

    elif nivel == 7:
        razon = just if just and just not in ("","nan") else "información insuficiente"
        return (
            f"Nivel 7. No georreferenciado. Razón: {razon}. "
            f"Sin coordenadas asignadas según protocolo SiB Colombia / Instituto Humboldt."
        )

    return ""


# ─── Hoja de resumen ───────────────────────────────────────────
def crear_hoja_resumen(ws, df, idioma):
    cols = nombres_columnas(idioma)

    def titulo_seccion(row, texto):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=texto)
        c.font = Font(bold=True, size=11, name="Calibri", color="FFFFFF")
        c.fill = AZUL_DARK
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20

    def dato(row, col, valor, bold=False, fill=None, align="left"):
        c = ws.cell(row=row, column=col, value=valor)
        c.font = Font(bold=bold, size=10, name="Calibri")
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if fill: c.fill = fill
        return c

    # Título principal
    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1,
                value="Resultados — Verificador de Georreferenciación SiB Colombia")
    c.font = Font(bold=True, size=13, name="Calibri", color="FFFFFF")
    c.fill = AZUL_DARK
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    fila = 3

    # ── 1. Leyenda ────────────────────────────────────────────
    titulo_seccion(fila, "QUE SIGNIFICAN LOS COLORES Y RESULTADOS")
    fila += 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
    c = ws.cell(row=fila, column=1, value=
        "Los colores en la hoja Registros indican el resultado de la validacion espacial de cada coordenada.")
    c.font = Font(italic=True, size=10, name="Calibri", color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    fila += 1

    leyenda = [
        ("Verde",   "OK",       "Coordenada válida y dentro del municipio reportado"),
        ("Amarillo","Revisar",  "Advertencia: coordenada en municipio vecino o incertidumbre alta"),
        ("Rojo",    "Error",    "Error: coordenada fuera de Colombia o dato crítico faltante"),
        ("Gris",    "—",        "Sin coordenadas / Nivel 7 — no georreferenciado"),
        ("Blanco",  "vacío",    "Registro sin validación espacial (sin coordenadas originales)"),
    ]
    leyenda_indicadores = [
        ("[✓]",  "Resultado correcto o campo completo"),
        ("[!]",  "Advertencia — requiere revision del curador"),
        ("[X]",  "Error — requiere correccion"),
        ("[A]",  "Campo que el codigo llena automaticamente"),
        ("[M]",  "Campo que el curador llena manualmente"),
    ]
    dummy = [  # dummy para no romper el loop original
    ]
    dato(fila, 1, "Color", bold=True, fill=AZUL_MED)
    dato(fila, 2, "Resultado", bold=True, fill=AZUL_MED)
    dato(fila, 3, "Significado", bold=True, fill=AZUL_MED)
    fila += 1

    colores_leyenda = {
        "Verde": VERDE, "Amarillo": AMARILLO,
        "Rojo": ROJO, "Gris": GRIS, "Blanco": None
    }
    for color_n, resultado, desc in leyenda:
        fill = colores_leyenda.get(color_n)
        dato(fila, 1, color_n, fill=fill)
        dato(fila, 2, resultado, fill=fill)
        ws.merge_cells(start_row=fila, start_column=3, end_row=fila, end_column=4)
        c = ws.cell(row=fila, column=3, value=desc)
        c.font = Font(size=10, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if fill: c.fill = fill
        fila += 1

    fila += 1

    # ── 2. Distribución por nivel ─────────────────────────────
    titulo_seccion(fila, "DISTRIBUCION POR NIVEL DE CALIDAD")
    fila += 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
    c = ws.cell(row=fila, column=1, value=
        "Clasificacion segun Tabla 9 del manual. Nivel 1 = coordenadas originales. "
        "Niveles 2-6 = sin coordenadas, georreferenciados por texto de localidad. "
        "Nivel 7 = no georreferenciable.")
    c.font = Font(italic=True, size=10, name="Calibri", color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[fila].height = 30
    fila += 1
    dato(fila, 1, "Nivel", bold=True, fill=AZUL_MED)
    dato(fila, 2, "Descripcion", bold=True, fill=AZUL_MED)
    dato(fila, 3, "Registros", bold=True, fill=AZUL_MED, align="center")
    dato(fila, 4, "%", bold=True, fill=AZUL_MED, align="center")
    fila += 1

    total = len(df)
    descripciones = {
        1: "Con coordenadas originales (GPS / campo)",
        2: "Con distancias u orientaciones — sin coords",
        3: "Entidades geograficas locales — sin coords",
        4: "Areas amplias o Sin datos con municipio",
        5: "Solo departamento reportado",
        6: "Solo pais reportado",
        7: "Informacion dudosa — NO georreferenciado",
    }
    colores_nivel = {
        1: None, 2: None, 3: None, 4: None,
        5: None, 6: None, 7: None
    }
    for nivel in range(1, 8):
        cnt  = int((df["Nivel_final"] == nivel).sum())
        pct  = f"{cnt/total*100:.1f}%" if cnt > 0 else "—"
        desc = descripciones.get(nivel, "")
        if cnt == 0:
            txt = f"Nivel {nivel}  —  Sin registros en esta base de datos"
            dato(fila, 1, txt, bold=False)
            dato(fila, 2, desc)
            dato(fila, 3, 0, align="center")
            dato(fila, 4, "—", align="center")
        else:
            dato(fila, 1, f"Nivel {nivel}", bold=True)
            dato(fila, 2, desc)
            dato(fila, 3, cnt, align="center")
            dato(fila, 4, pct, align="center")
        dato(fila, 4, pct,  fill=fill, align="center")
        fila += 1

    fila += 1

    # ── 3. Campos faltantes ───────────────────────────────────
    titulo_seccion(fila, "CAMPOS FALTANTES — Revisiones minimas (Tabla 12 del manual)")
    fila += 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
    c = ws.cell(row=fila, column=1, value=
        "Descripcion de los indicadores en la columna Estado:")
    c.font = Font(italic=True, size=10, name="Calibri", color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center")
    fila += 1

    F_VERDE   = PatternFill("solid", fgColor="C6EFCE")
    F_AMARILLO= PatternFill("solid", fgColor="FFEB9C")
    F_NARANJA = PatternFill("solid", fgColor="FCE4D6")
    leyenda_campos = [
        ("[✓] Completo",              F_VERDE,    "El campo ya tiene datos en todos los registros."),
        ("[A] Completar con proceso",  F_AMARILLO, "El codigo lo llena automaticamente al ejecutar todos los bloques (elevacion, incertidumbre, fuentes)."),
        ("[M] Llenar manualmente",     F_NARANJA,  "El curador o investigador debe completarlo a mano (quien georreferencio y en que fecha)."),
    ]
    for (ind, fill_l, desc_l) in leyenda_campos:
        dato(fila, 1, ind, fill=fill_l, bold=True)
        ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=4)
        c2 = ws.cell(row=fila, column=2, value=desc_l)
        c2.font = Font(size=10, name="Calibri")
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c2.fill = fill_l
        fila += 1
    fila += 1
    dato(fila, 1, "Campo", bold=True, fill=AZUL_MED)
    dato(fila, 2, "Vacios", bold=True, fill=AZUL_MED, align="center")
    dato(fila, 3, "Estado", bold=True, fill=AZUL_MED)
    fila += 1

    NARANJA      = PatternFill("solid", fgColor="FCE4D6")  # curador llena manual
    AMARILLO_INF = PatternFill("solid", fgColor="FFEB9C")  # código llena automático (pendiente)

    campos_revisar = [
        (cols["datum"],         "Datum",                             "codigo"),
        (cols["comentarios"],   "Comentarios de la georreferenciacion","codigo"),
        (cols["fuentes"],       "Fuentes de georreferenciacion",     "codigo"),
        (cols["protocolo"],     "Protocolo de georreferenciacion",   "codigo"),
        (cols["incertidumbre"], "Incertidumbre de coordenadas (m)",  "codigo"),
        (cols["por"],           "Georreferenciado por",              "curador"),
        (cols["fecha"],         "Fecha de georreferenciacion",       "curador"),
    ]
    for col_nombre, etiqueta, quien in campos_revisar:
        if col_nombre in df.columns:
            vacios = int(
                df[col_nombre].isna().sum() +
                (df[col_nombre].astype(str).str.strip() == "").sum()
            )
            if vacios == 0:
                fill = VERDE
                estado = "[✓] Completo"
            elif quien == "curador":
                fill = PatternFill('solid', fgColor='FCE4D6')
                estado = f"[M] Llenar manualmente - {vacios} pendientes"
            else:
                fill = PatternFill('solid', fgColor='FFEB9C')
                estado = f"[A] Completar con el proceso - {vacios} pendientes"
            dato(fila, 1, etiqueta, fill=fill)
            dato(fila, 2, vacios,   fill=fill, align="center")
            dato(fila, 3, estado,   fill=fill)
            fila += 1

    fila += 1

    # ── 4. Validación espacial ────────────────────────────────
    if "validacion_b2" in df.columns or "Resultado validación espacial" in df.columns:
        titulo_seccion(fila, "VALIDACION ESPACIAL — Registros con coordenadas (Niveles 1-6)")
        fila += 1
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
        c = ws.cell(row=fila, column=1, value=
            "Verifica si la coordenada asignada cae dentro del municipio reportado "
            "(seccion 3.6.1 del manual). OK = dentro | Revisar = municipio vecino | "
            "Error = fuera de Colombia.")
        c.font = Font(italic=True, size=10, name="Calibri", color="595959")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        fila += 1
        col_val = "Resultado validación espacial" if "Resultado validación espacial" in df.columns else "validacion_b2"
        # Solo nivel 1 tiene validación espacial real (coordenadas del colector)
        # Niveles 2-6 tienen centroide asignado — se muestran en desglose aparte
        val_counts = df[df["Nivel_final"] == 1][col_val].value_counts()
        desc_val = {
            "✅": "Punto dentro del municipio reportado",
            "OK": "Punto dentro del municipio reportado",
            "⚠":  "Punto en municipio vecino — revisar",
            "Revisar": "Punto en municipio vecino — revisar",
            "❌": "Punto fuera de Colombia o error grave",
            "Error": "Punto fuera de Colombia o error grave",
            "🔵": "Sin coordenadas convertibles",
            "Sin coords": "Sin coordenadas convertibles",
        }
        for estado, cnt in val_counts.items():
            fill = VERDE if estado in ("✅","OK") else (
                   AMARILLO if estado in ("⚠","Revisar") else (
                   ROJO if estado in ("❌","Error") else None))
            desc = desc_val.get(str(estado), str(estado))
            dato(fila, 1, texto_validacion(estado), fill=fill, bold=True)
            dato(fila, 2, desc, fill=fill)
            dato(fila, 3, int(cnt), fill=fill, align="center")
            fila += 1
        # Registros sin validación espacial — desglose por nivel
        desglose = {
            2: ("Nivel 2", "Sin coords originales — coordenada inferida desde topónimos OSM o centroide del municipio"),
            3: ("Nivel 3", "Sin coords originales — coordenada asignada al centro geográfico del municipio reportado"),
            4: ("Nivel 4", "Sin coords originales — coordenada asignada al centro geográfico del municipio reportado"),
            5: ("Nivel 5", "Sin coords originales — coordenada asignada al centro geográfico del departamento reportado"),
            6: ("Nivel 6", "Sin coords originales — coordenada asignada al centro geográfico de Colombia"),
        }
        for nv, (etiqueta, descripcion) in desglose.items():
            cnt_nv = int((df["Nivel_final"] == nv).sum())
            if cnt_nv > 0:
                dato(fila, 1, etiqueta, bold=False)
                dato(fila, 2, descripcion)
                dato(fila, 3, cnt_nv, align="center")
                fila += 1
        fila += 1

    # ── 5. Totales ────────────────────────────────────────────
    titulo_seccion(fila, "TOTALES")
    fila += 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
    c = ws.cell(row=fila, column=1, value="Resumen general del proceso sobre los 264 registros.")
    c.font = Font(italic=True, size=10, name="Calibri", color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center")
    fila += 1
    datos_totales = [
        ("Total registros procesados",                total),
        ("Nivel 1 — con coordenadas originales",      int((df["Nivel_final"]==1).sum())),
        ("Niveles 2-6 — georreferenciados por texto", int(df["Nivel_final"].isin([2,3,4,5,6]).sum())),
        ("Nivel 7 — no georreferenciables",           int((df["Nivel_final"]==7).sum())),
    ]
    for nombre, cnt in datos_totales:
        pct = f"{cnt/total*100:.1f}%" if total > 0 else "—"
        dato(fila, 1, nombre)
        dato(fila, 2, cnt, align="center", bold=True)
        dato(fila, 3, pct, align="center")
        fila += 1

    # Anchos
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10


# ─── Función principal ─────────────────────────────────────────
def aplicar_bloque10(df, ruta_salida=None, idioma=None):
    if idioma is None:
        idioma = detectar_idioma(df)

    print(f"  Preparando Excel final — idioma: {idioma} — {len(df)} registros")
    df = preparar_dataframe(df, idioma)
    cols = nombres_columnas(idioma)

    # Guardar copia para resumen (con nombres originales de columnas internas)
    df_stats = df.copy()

    # Columnas internas a ocultar
    COLS_INTERNAS = {
        "lat_wgs84", "lon_wgs84", "fuente_conversion",
        "municipio_detectado", "depto_detectado", "mensaje_b2",
    }

    # Renombrar columnas visibles
    renombrar = {
        "Nivel_inicial":       "Nivel de calidad inicial",
        "Nivel_final":         "Nivel de calidad final",
        "Justificacion_nivel": "Justificación del nivel",
        "validacion_b2":       "Resultado validación espacial",
    }

    # Columna Origen — valores legibles, sin duplicados
    if "Origen" in df.columns:
        df["Origen"] = df["Origen"].replace({
            "180_con_coordenadas": "Con coordenadas",
            "84_sin_coordenadas":  "Sin coordenadas",
        })
        df = df.loc[:, ~df.columns.duplicated()]

    # Convertir emojis en columna validacion a texto limpio
    if "validacion_b2" in df.columns:
        df["validacion_b2"] = df["validacion_b2"].apply(texto_validacion)

    df = df.rename(columns=renombrar)

    # Seleccionar columnas a exportar
    cols_exportar = [c for c in df.columns if c not in COLS_INTERNAS]
    df_export = df[cols_exportar].copy()
    cols_nuevas = COLUMNAS_NUEVAS_ES if idioma == "es" else COLUMNAS_NUEVAS_DWC

    # Convertir todas las columnas a tipos compatibles con Excel
    # Evita "Invalid value for dtype str" cuando hay floats mezclados con strings
    import numpy as np
    for col in df_export.columns:
        if df_export[col].dtype == object:
            # Columna de texto — asegurar que todos los valores sean string
            df_export[col] = df_export[col].apply(
                lambda x: "" if (x is None or (isinstance(x, float) and np.isnan(x)))
                else str(x)
            )
        elif df_export[col].dtype in [np.float64, np.float32]:
            # Columna numérica float — convertir a string con punto
            df_export[col] = df_export[col].apply(
                lambda x: f"{x:.6f}" if pd.notna(x) else ""
            )

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        df_export.to_excel(writer, sheet_name="Registros", index=False)
        wb = writer.book
        ws = writer.sheets["Registros"]

        # Anchos de columna según tipo de contenido
        ANCHOS = {
            # Columnas de texto largo — wrap text
            cols["comentarios"]:          45,
            cols["fuentes"]:              45,
            cols["protocolo"]:            45,
            "Justificación del nivel":    40,
            # Coordenadas — ancho exacto para decimal
            cols["lat_orig"]:             18,
            cols["lon_orig"]:             18,
            cols["lat_final"]:            22,
            cols["lon_final"]:            22,
            # Columnas de resultado
            "Nivel de calidad inicial":   20,
            "Nivel de calidad final":     20,
            "Resultado validación espacial": 22,
            # Geografía
            cols["municipio"]:            18,
            cols["departamento"]:         18,
            cols["datum"]:                16,
            "Origen":                     18,
        }

        for col_idx, col_name in enumerate(df_export.columns, 1):
            c = ws.cell(row=1, column=col_idx)
            es_nueva = col_name in cols_nuevas
            c.fill = VERDE_COLS if es_nueva else AZUL_DARK
            c.font = Font(
                bold=True, name="Calibri", size=10,
                color="000000" if es_nueva else "FFFFFF"
            )
            c.alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)
            # Ancho personalizado o default 16
            ancho = ANCHOS.get(col_name, 16)
            ws.column_dimensions[get_column_letter(col_idx)].width = ancho

        ws.row_dimensions[1].height = 35
        ws.freeze_panes = "A2"

        # Coordenadas como texto con punto fijo — evita problema de separador regional
        for col_name in df_export.columns:
            if "georreferenciada" in str(col_name).lower() or \
               col_name in ("decimalLatitude_georef", "decimalLongitude_georef"):
                col_idx = df_export.columns.get_loc(col_name) + 1
                for row_idx in range(2, len(df_export) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and str(cell.value) not in ("", "nan", "None"):
                        try:
                            cell.value = f"{float(cell.value):.6f}"
                        except (ValueError, TypeError):
                            pass

        # Colorear columnas de resultado y activar wrap text en celdas de datos
        cols_colorear = {
            "Nivel de calidad inicial",
            "Nivel de calidad final",
            "Justificación del nivel",
            "Resultado validación espacial",
            cols["lat_final"],
            cols["lon_final"],
            cols["comentarios"],
            cols["fuentes"],
            cols["datum"],
        }
        # Columnas de texto largo que necesitan wrap text
        cols_wrap = {
            cols["comentarios"],
            cols["fuentes"],
            cols["protocolo"],
            "Justificación del nivel",
        }

        for row_idx, (_, row) in enumerate(df_export.iterrows(), 2):
            fill = color_por_fila(row)
            ws.row_dimensions[row_idx].height = 45  # altura fija para wrap text
            for col_idx, col_name in enumerate(df_export.columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Wrap text en columnas de texto largo
                if col_name in cols_wrap:
                    cell.alignment = Alignment(
                        wrap_text=True, vertical="top", horizontal="left"
                    )
                else:
                    cell.alignment = Alignment(
                        wrap_text=False, vertical="center", horizontal="left"
                    )
                # Color solo en columnas de resultado
                if col_name in cols_colorear:
                    cell.fill = fill

        # Hoja Resumen
        wb.create_sheet("Resumen")
        crear_hoja_resumen(wb["Resumen"], df_stats, idioma)
        wb.move_sheet("Resumen", offset=-len(wb.sheetnames)+1)

    output.seek(0)
    datos = output.getvalue()

    if ruta_salida:
        with open(ruta_salida, "wb") as f:
            f.write(datos)
        print(f"  Guardado: {ruta_salida}")
        return ruta_salida
    return datos


# ─── Prueba ────────────────────────────────────────────────────
